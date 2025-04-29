import openpyxl
import traceback
import numpy as np
import pandas as pd
from tkinter import filedialog, messagebox
from ui.dialogs import LoadFileDialog
from core.app_state import clear_plot
from core.calculate_baseline import calculate_baseline

def load_file(app):
    """
    Load data from an Excel file
    
    Args:
        app: The main application instance
    
    Returns:
        bool: If the file is successfully loaded, return True, otherwise return False
    """
    try:

        # Create and display the input dialog, using the last input as the default value
        load_file_dialog = LoadFileDialog(
            app,
            default_sheet_name=app.last_sheet_name,
            default_x_col=app.last_x_col,
            default_y_col=app.last_y_col,
            default_RFP_col=app.last_RFP_col
        )

        try:
            # Bind events before displaying the dialog
            def handle_destroy(event):
                if event.widget == app:
                    load_file_dialog.user_cancelled = True
                    load_file_dialog.destroy()
            
            app.bind('<Destroy>', handle_destroy)
            app.wait_window(load_file_dialog)  # Wait for the dialog to close
            app.unbind('<Destroy>')
        except Exception as e:
            # Ignore binding related errors when the window is closed
            return False

        # Check if the user has cancelled the operation
        if load_file_dialog.user_cancelled:
            return False

        # Get input from the dialog
        sheet_name = load_file_dialog.sheet_name
        x_col = load_file_dialog.x_col
        y_col = load_file_dialog.y_col
        RFP_col = load_file_dialog.RFP_col
        RFP_smoothing_window_size = load_file_dialog.RFP_smoothing_window_size
        convert_to_dr_r = load_file_dialog.convert_to_dr_r
        app.convert_to_df_f = load_file_dialog.convert_to_df_f     
        app.evoked_status = load_file_dialog.evoked_status

        if not sheet_name or not x_col or not y_col:
            messagebox.showwarning(title="Warning", message="All fields must be filled out.")
            return False

        # Update the last used input values
        app.last_sheet_name = sheet_name
        app.last_x_col = x_col
        app.last_y_col = y_col

        # Use the file dialog to select a file
        file_path = filedialog.askopenfilename()

        if not file_path:
            messagebox.showwarning(title="Warning", message="No file selected.")
            return False

        # Clear the previous chart
        clear_plot(app, reset_data=True)
        
        try:
            # Set the progress bar to 0
            app.progress_bar.set(0)
            
            # Use the read-only mode to load the Excel file
            app.progress_bar.set(0.1)
            
            # Use the read-only mode to open the Excel file
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            app.progress_bar.set(0.2)
            
            # Get the specified sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                messagebox.showerror(title="Error", message=f"Sheet '{sheet_name}' not found in the workbook.")
                app.progress_bar.set(0)
                return False
            
            # Get the header row
            header_row = next(ws.rows)
            header = [cell.value for cell in header_row]
            
            # Convert the column names to strings
            header = [str(col) for col in header]
            
            # Find the indices of the x and y columns
            try:
                x_idx = header.index(x_col)
            except ValueError:
                error_msg = f"Column '{x_col}' not found in the sheet."
                messagebox.showerror(title="Error", message=error_msg)
                app.progress_bar.set(0)
                return False
                
            try:
                y_idx = header.index(y_col)
            except ValueError:
                error_msg = f"Column '{y_col}' not found in the sheet."
                messagebox.showerror(title="Error", message=error_msg)
                app.progress_bar.set(0)
                return False
                
            # if DR/R mode, find the index of the RFP column
            if convert_to_dr_r and RFP_col:
                try:
                    rfp_idx = header.index(RFP_col)
                except ValueError:
                    error_msg = f"Column '{RFP_col}' not found in the sheet."
                    messagebox.showerror(title="Error", message=error_msg)
                    app.progress_bar.set(0)
                    return False
            
            # Prepare data lists
            app.time = []
            app.df_f = []
            rfp_values = [] if convert_to_dr_r else None

            # Get the total number of rows estimate (cannot directly get the number of rows in read-only mode)
            # Using ws.max_row may be inaccurate, but can be used as a reference for the progress bar
            total_rows_estimate = ws.max_row - 1  # Subtract the header row
            
            # Read data rows
            row_count = 0
            for row in ws.rows:
                row_count += 1
                if row_count == 1:  # Skip the header row
                    continue
                
                # Get the x and y values
                try:
                    x_val = row[x_idx].value
                    y_val = row[y_idx].value
                    
                    # if DR/R mode, get the RFP value
                    if convert_to_dr_r:
                        rfp_val = row[rfp_idx].value
                        if x_val is not None and y_val is not None and rfp_val is not None and float(rfp_val) != 0:
                            app.time.append(float(x_val))
                            app.df_f.append(float(y_val))
                            rfp_values.append(float(rfp_val))
                    else:
                        if x_val is not None and y_val is not None:
                            app.time.append(float(x_val))
                            app.df_f.append(float(y_val))
                except (IndexError, TypeError, ValueError) as e:
                    # Skip problematic rows
                    continue
                
                # Update the progress bar
                if row_count % 1000 == 0:  # Update the progress bar every 1000 rows
                    progress = min(0.2 + 0.7 * row_count / total_rows_estimate, 0.9)
                    app.progress_bar.set(progress)
                    app.update()  # Force update GUI
            
            # Close the workbook
            wb.close()
            
            # Check if the data was successfully read
            if not app.time or not app.df_f:
                messagebox.showerror(title="Error", message="No valid data found in the selected columns.")
                app.progress_bar.set(0)
                return False
            
            # Convert to pandas Series
            app.time = pd.Series(app.time)
            app.df_f = pd.Series(app.df_f)
            
            # handle DR/R case
            if convert_to_dr_r:
                rfp_values = pd.Series(rfp_values)

                # if RFP smoothing window size is not empty, perform smoothing
                if RFP_smoothing_window_size and int(RFP_smoothing_window_size) > 1:
                    rfp_values = rfp_values.rolling(window=int(RFP_smoothing_window_size), center=True, min_periods=1).mean()

                app.df_f = app.df_f / rfp_values
                calculate_baseline(app)
                
                # calculate final DR/R
                app.df_f = (app.df_f - app.baseline_values) / app.baseline_values
                app.df_f = app.df_f.replace([np.inf, -np.inf], np.nan)
                app.df_f = app.df_f.fillna(0)
                app.progress_bar.set(0.98)
            # handle DF/F case
            elif app.convert_to_df_f and app.df_f.mean() > 3:
                # calculate baseline
                calculate_baseline(app)
                
                # calculate DF/F
                app.df_f = (app.df_f - app.baseline_values) / app.baseline_values
                app.df_f = app.df_f.replace([np.inf, -np.inf], np.nan)
                app.df_f = app.df_f.fillna(0)
                app.progress_bar.set(0.98)
            
            if 0 <= app.df_f.mean() <= 3:
                app.convert_to_df_f = True

            calculate_baseline(app)

            # Draw the chart
            app.ax.clear()
            app.ax.plot(app.time, app.df_f, color='black')
            app.ax.set_ylim(np.min(app.df_f), np.max(app.df_f))
            app.ax.grid(True)
            app.canvas.draw()
            
            # Complete
            app.progress_bar.set(1.0)
            
            # Reset the progress bar after a delay
            app.after(500, lambda: app.progress_bar.set(0))
            
            return True
            
        except Exception as e:
            app.progress_bar.set(0)
            error_message = f"Error loading file: {str(e)}"
            print(error_message)
            print(traceback.format_exc()) 
            messagebox.showerror(title="Error", message=error_message)
            return False
    except Exception as e:
        error_message = f"Error in load_file function: {str(e)}"
        print(error_message)
        print(traceback.format_exc()) 
        return False  # Ignore errors when closing